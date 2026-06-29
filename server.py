from flask import Flask, request, send_file, send_from_directory
from openpyxl import load_workbook
from datetime import datetime
import io
import os
import traceback

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)

TEMPLATE = os.path.join(BASE_DIR, "PLANILLA DE RENDICION v1.0.xlsx")
INDEX_HTML = os.path.join(BASE_DIR, "index.html")

RANGES = {
    "GIGANTE": (5, 7),
    "OBRAS": (9, 11),
    "LM": (13, 18),
}

MEDIO_TO_COL = {
    "EFECTIVO": "E",
    "TRANSF.": "F",
    "CHEQUES": "H",
    "ECHEQ": "I",
    "RETENCIONES": "K",
    "AJUSTE": "J",
}

def find_next_row(ws, start, end):
    for r in range(start, end + 1):
        if ws[f"D{r}"].value in (None, ""):
            return r
    return None

def add_number(ws, cell_addr, value):
    current = ws[cell_addr].value
    try:
        current_num = float(current) if current not in (None, "") else 0.0
    except Exception:
        current_num = 0.0
    ws[cell_addr].value = current_num + float(value)

def build_excel_from_clients(clients):
    if not os.path.exists(TEMPLATE):
        raise FileNotFoundError(f"No encuentro la plantilla: {TEMPLATE}")

    wb = load_workbook(TEMPLATE)
    ws = wb["Rendición"] if "Rendición" in wb.sheetnames else wb.active
    ws["C3"].value = datetime.now().strftime("%d/%m/%Y")

    for c in clients:
        modal = c.get("modal")
        cli = (c.get("cli") or "").strip().upper()
        items = c.get("items", [])

        if modal not in RANGES or not cli or not items:
            continue

        start, end = RANGES[modal]
        row = find_next_row(ws, start, end)
        if row is None:
            continue

        ws[f"D{row}"].value = cli

        for it in items:
            med = it.get("med")
            imp = it.get("imp")

            if med not in MEDIO_TO_COL:
                continue
            try:
                imp = float(imp)
            except Exception:
                continue

            col = MEDIO_TO_COL[med]
            add_number(ws, f"{col}{row}", imp)

    filename = f"rendicion-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue(), filename

@app.get("/")
def home():
    if not os.path.exists(INDEX_HTML):
        return (
            "No encuentro index.html en la misma carpeta que server.py.\n"
            "Poné el archivo como: <carpeta>/index.html",
            404,
        )
    return send_from_directory(BASE_DIR, "index.html")

@app.get("/health")
def health():
    return "OK", 200

@app.post("/generar")
def generar():
    inline = request.args.get("inline") == "1"
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return "JSON inválido", 400

        clients = data.get("clients", [])
        if not clients:
            return "Sin datos", 400

        excel_bytes, filename = build_excel_from_clients(clients)
        bio = io.BytesIO(excel_bytes)

        response = send_file(
            bio,
            as_attachment=not inline,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            max_age=0,
        )
        if inline:
            response.headers["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    except Exception:
        return "ERROR:\n" + traceback.format_exc(), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run("0.0.0.0", port, debug=False)
